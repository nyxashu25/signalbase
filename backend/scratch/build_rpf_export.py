import json
import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

SRC = r"C:\Users\digig\Downloads\RPF_Format.xlsx"
CONTACTS_JSON = r"C:\Users\digig\Downloads\code\backend\contacts_export.json"
OUT = r"C:\Users\digig\Downloads\RPF_SignalBase_Export.xlsx"

HEADERS = [
    "Date", "First Name", "Last Name", "Job Title", "Email ID", "Domain",
    "Department", "Seniority", "Company Name", "Industry Type", "TelephoneNo",
    "Alternative No.", "Address", "City", "Zip Code", "State", "Country",
    "Emp Size", "Revenue", "Prospect Linkedin profile Link", "Company Linkedin profile Link",
]

# SignalBase's industry values don't literally appear in the RPF picklist
# (which is a general-business enum, not a tech-vertical one) — mapped to
# the closest matching RPF category. Documented in the Notes sheet.
INDUSTRY_MAP = {
    "SaaS": "Software and IT Services",
    "Fintech": "Finance",
    "Healthcare": "Healthcare",
    "E-commerce": "Retail",
    "Cybersecurity": "Software and IT Services",
    "Logistics": "Transportation and Logistics",
    "Marketing": "Media and Communications",
    "EdTech": "Education",
}

STATE_MAP = {
    "TX": "Texas", "CA": "California", "MA": "Massachusetts", "CO": "Colorado",
    "IL": "Illinois", "WA": "Washington", "NY": "New York",
}

EMP_SIZE_BUCKETS = [
    (0, 1, "0-1"), (2, 10, "2-10"), (11, 50, "11-50"), (51, 200, "51-200"),
    (201, 500, "201-500"), (501, 1000, "501-1000"), (1001, 5000, "1001-5000"),
    (5001, 10000, "5001-10000"), (10001, float("inf"), "10001+"),
]


def emp_size_bucket(headcount_min, headcount_max):
    if headcount_min is None or headcount_max is None:
        return None
    midpoint = (headcount_min + headcount_max) / 2
    for lo, hi, label in EMP_SIZE_BUCKETS:
        if lo <= midpoint <= hi:
            return label
    return None


def parse_location(location):
    if not location:
        return None, None, "United States"
    city, _, region = location.partition(",")
    region = region.strip()
    city = city.strip()
    if region == "ON":
        return city, "Ontario", "Canada"
    return city, STATE_MAP.get(region, region), "United States"


def main():
    with open(CONTACTS_JSON, encoding="utf-8") as f:
        contacts = json.load(f)

    wb = openpyxl.load_workbook(SRC)
    reference_sheet = wb["RPF_Format"]
    reference_sheet.title = "RPF Format (reference)"

    ws = wb.create_sheet("Data", 0)
    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    body_font = Font(name="Arial", size=10)

    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"

    today = datetime.date.today().strftime("%d-%m-%Y")
    unmapped_industries = set()

    for row_idx, contact in enumerate(contacts, start=2):
        company = contact["company"]
        industry = company.get("industry")
        rpf_industry = INDUSTRY_MAP.get(industry, industry)
        if industry and industry not in INDUSTRY_MAP:
            unmapped_industries.add(industry)

        city, state, country = parse_location(company.get("location"))
        emp_size = emp_size_bucket(company.get("headcountMin"), company.get("headcountMax"))

        values = [
            today,
            contact.get("firstName"),
            contact.get("lastName"),
            contact.get("title"),
            contact.get("email"),  # blank = not yet found/revealed, matches live DB state
            company.get("domain"),
            contact.get("department"),
            contact.get("seniority"),
            company.get("name"),
            rpf_industry,
            None,  # TelephoneNo — not collected by SignalBase's data model
            None,  # Alternative No.
            None,  # Address — not collected
            city,
            None,  # Zip Code — not collected
            state,
            country,
            emp_size,
            None,  # Revenue — not collected
            None,  # Prospect LinkedIn — not collected
            None,  # Company LinkedIn — not collected
        ]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.font = body_font

    widths = [12, 14, 14, 26, 26, 20, 16, 12, 20, 22, 12, 14, 20, 14, 10, 16, 14, 12, 12, 30, 30]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Notes sheet documenting every assumption made mapping SignalBase's
    # schema onto the RPF format, per the source data's actual availability.
    notes = wb.create_sheet("Notes")
    notes["A1"] = "Export notes — SignalBase demo data mapped to RPF format"
    notes["A1"].font = Font(name="Arial", bold=True, size=12)
    lines = [
        "",
        f"Generated: {datetime.date.today().isoformat()}",
        f"Source: SignalBase local Postgres dev database (40 companies, {len(contacts)} contacts)",
        "",
        "Fields left blank — not part of SignalBase's current data model:",
        "  TelephoneNo, Alternative No., Address, Zip Code, Revenue,",
        "  Prospect Linkedin profile Link, Company Linkedin profile Link",
        "",
        "Email ID is blank for every row: none of these contacts have had their email",
        "revealed yet in SignalBase (the reveal action is credit-gated). This is the",
        "true current state of the seed data, not a mapping gap.",
        "",
        "Industry Type — SignalBase's industries are tech-vertical labels (SaaS, Fintech,",
        "EdTech, Cybersecurity) that don't appear in the RPF picklist, which is a general-",
        "business enum. Mapped to the closest RPF category:",
    ]
    for k, v in INDUSTRY_MAP.items():
        lines.append(f"    {k} -> {v}")
    lines += [
        "",
        "State — SignalBase stores 'City, ST' (e.g. 'Austin, TX'); expanded to the RPF's",
        "full state name. One location, Toronto ON, is not a US state — mapped to",
        "State=Ontario, Country=Canada rather than forced into the US-only picklist.",
        "",
        "Emp Size — SignalBase stores a headcount min/max range (e.g. 150-350); bucketed",
        "into the RPF's single Emp Size band using the range midpoint.",
        "",
        "The original template's picklist/reference sheet is preserved unchanged as",
        "'RPF Format (reference)'.",
    ]
    for i, line in enumerate(lines, start=2):
        notes.cell(row=i, column=1, value=line).font = Font(name="Arial", size=10)
    notes.column_dimensions["A"].width = 90

    if unmapped_industries:
        print("WARNING unmapped industries:", unmapped_industries)

    wb.save(OUT)
    print(f"Wrote {len(contacts)} rows to {OUT}")


if __name__ == "__main__":
    main()
